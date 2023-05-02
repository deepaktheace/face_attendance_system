import openpyxl
import pandas as pd
from datetime import date
from os import remove
from pickle import load

df = pd.read_excel(f"attendance sheets/{date.today().day}-{date.today().month}-{date.today().year}.xlsx",
                   usecols=[1,8], sheet_name=f"{date.today().day}-{date.today().month}")
wb = openpyxl.load_workbook(f"attendance sheets/{date.today().day}-{date.today().month}-{date.today().year}.xlsx")

cur_sheet = wb[f"{date.today().day}-{date.today().month}"]

roll_num_list = open('ids.p','rb')
roll_number,time = load(roll_num_list)
roll_num_list.close()
roll_number = int(roll_number)
print(f"updaing {roll_number}")

for i,index in enumerate(df.iterrows()):
    if index[1]['ID'] == roll_number:
      cur_sheet.cell(i+2,9).value = "YES"
      cur_sheet.cell(i+2,10).value = time


print("updated")

remove(f"attendance sheets/{date.today().day}-{date.today().month}-{date.today().year}.xlsx")

wb.save(f"attendance sheets/{date.today().day}-{date.today().month}-{date.today().year}.xlsx")